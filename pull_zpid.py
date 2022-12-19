from aiohttp import ClientSession, TCPConnector
import asyncio
import pypeln as pl
import asyncio
from aiohttp import ClientSession
from time import time
from bs4 import BeautifulSoup
import openpyxl
import pandas as pd
import pandas as pd


excelname = 'enter .xlsx file here'


#Open xlsx file containing addresses and create a new row for zpid
df = pd.read_excel(excelname)
wb = openpyxl.load_workbook(excelname)
sheet = wb.active
sheet.cell(row=1, column=6).value = "zpid"

#necessary lists
data = []
url_list = []


#helper for formatting urls 
def removeat(string):
    return string.replace(" ", "+")

for val in df.values:
    data.append(val)

#Create a list of urls to be used in the requests
for idx, d in enumerate(data):
    f = {}
    urlparam = str(d[4]) + " zillow home details"
    urlparam = removeat(urlparam)
    f["url"] = 'https://www.bing.com/search?q=' + urlparam
    f["index"] = idx + 2
    url_list.append(f)


start = time()

limit = 100
async def main():

    async with ClientSession(connector=TCPConnector(limit=0)) as session:

        async def fetch(url):
            async with session.get(url["url"]) as response:
                #Gather responses from urls asynchronously
                content = await response.text()
                print("--------------------------------------------")
                #Parse response and find zpid
                soup = BeautifulSoup(content, "html.parser")
                zpid = str(soup.select("a[href*=_zpid]")).split('_zpid')[0].split('/')[-1]
                print(zpid)
                print("index is " + str(url["index"]))
                sheet.cell(row=url["index"], column=6).value = zpid
                return(zpid)

        await pl.task.each(
            fetch, url_list, workers=limit,
        )


asyncio.run(main())
wb.save(excelname)
print("done")
print(f'Time taken: {time() - start}')
